
# -*- coding: utf-8 -*-
"""
Created on Mon Aug 24 14:51:18 2020

@author: rog
"""

import pandas as pd
import tkinter as tk
from tkinter import *
#from tkinter import ttk
from tkinter.ttk import *
from PIL import Image, ImageTk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename, asksaveasfile
from tkinter import Tk, Frame, Menu   
from tkinter import Button
import numpy as np
import sys 
import os
sys.path.append(os.path.abspath('Func'+ os.sep))
from tkinter import messagebox
from scipy.signal import butter, lfilter
import scipy.signal as signal
import statsmodels.api as sm
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from statsmodels.tsa.statespace.sarimax import SARIMAX
from pmdarima import auto_arima
import warnings
from sklearn.metrics import mean_squared_error
from statsmodels.tools.eval_measures import rmse
import scipy.stats



class Window(Frame):#(tk.Frame):
    def __init__(self, master):
        Frame.__init__(self, master)   
        frame = Frame(master)
        frame.pack()
        self.init_window()
        root = tk.Tk()
        frame = tk.Frame(root)
        frame.grid()
        
        self.imagecenit()
        self.Next()
        self.OpenFile()
        self.plotserieinicial()
        self.filtro()
        self.crosscorrel()
        self.correl()
        self.analisis_regresion()
        self.analisis_parametros_ARIMA()
        self.test_resultados_modelo_arima()
        self.mostrar_pronostico()
        self.graficar_todo()
        self.graficar_una_serie()

    def imagecenit(self):
        image = Image.open("LogoGrupoCENIT_2.png")
        photo = ImageTk.PhotoImage(image)
        self.canvas = tk.Canvas(self, height=768, width=1024)
        self.canvas.create_image(20, 10, anchor='nw', image=photo)
        self.canvas.grid(row=2, column=1)
        root.mainloop()

    def init_window(self):
        self.pack(fill=BOTH, expand=1)
        menu = Menu(self.master)
        self.master.config(menu=menu)
        ##############################################
        AnalisisEspectral = Menu(menu, tearoff=0)
        AnalisisEspectral.add_command(label="Open File", command=self.OpenFile)
        AnalisisEspectral.add_command(label="Plot Serie Inicial", command=self.plotserieinicial)
        AnalisisEspectral.add_command(label="Análisis Espectral", command=self.analisisespectral)
        AnalisisEspectral.add_command(label="Filtrar", command=self.filtro)
        menu.add_cascade(label="Análisis Espectral", menu=AnalisisEspectral)
        #############################################
        Correlacion_Cruzada = Menu(menu, tearoff=0)
        Correlacion_Cruzada.add_command(label="Análisis de Correlación Cruzada", command=self.crosscorrel)
        Correlacion_Cruzada.add_command(label="Correlación", command=self.correl)
        Correlacion_Cruzada.add_command(label="Análisis de Regresión", command=self.analisis_regresion)
        Correlacion_Cruzada.add_command(label="Exit", command=self.close_win)
        menu.add_cascade(label="Correlacion_Cruzada", menu=Correlacion_Cruzada)
        #############################################
        #############################################
        Prediccion_arima = Menu(menu, tearoff=0)
        Prediccion_arima.add_command(label="Análisis parámetros_ARIMA", command=self.analisis_parametros_ARIMA)
        Prediccion_arima.add_command(label="Resultados Test ARIMA", command=self.test_resultados_modelo_arima)
        Prediccion_arima.add_command(label="Mostrar Pronostico", command=self.mostrar_pronostico)
        Prediccion_arima.add_command(label="Graficar todo", command=self.graficar_todo)
        Prediccion_arima.add_command(label="Graficar una serie", command=self.graficar_una_serie)
        menu.add_cascade(label="Predicción_arima", menu=Prediccion_arima)
        #############################################
        helpmenu = Menu(menu, tearoff=0)
        helpmenu.add_command(label="Ayuda", command=self.show_help)
        helpmenu.add_command(label="About", command=self.about, foreground='blue',font = ('bold',10))#Verdana
        menu.add_cascade(label="Help", menu=helpmenu)
        ##################################################
        self.imagecenit()
        #####################################################################################################



    def showImg(self):
        load = Image.open(" ")
        render = ImageTk.PhotoImage(load)
        img = Label(self, image=render)
        img.image = render
        img.place(x=0, y=0)


    def show_help(self):
        win = Toplevel(root)
        win.geometry("300x300")
        lab = Label(win, text="Hey there good lookin!",foreground='blue',font = ('bold',10))
        lab.pack()
        

    def close_win(self):
        root.destroy()
    

    def about(self):
        win = Toplevel(root)
        win.geometry("300x150")
        lab = Label(win, text="Tratamiento de Datos\n\n Versión del programa 1.0\n\n Realizado por grupo CENIT:\n "
                              "Igor Malikov y Nancy Villegas",foreground='blue',font = ('bold',10),justify=CENTER)
        lab.pack()


    def Next(self, relx, rely):
        var = tk.IntVar()
        btn = tk.Button(root, width=14, text='Next', font="Times 12 bold", fg='red', command=lambda: var.set(1))
        btn.place(relx=relx, rely=rely, anchor="c")
        btn.wait_variable(var)
        btn.place_forget()
        self.canvas.delete('all')


    def text_canvas_StringVar(self,x1,y1,text1,text2,ancho,x2,y2):
        global nombr_variable_str
        self.canvas.create_text(x1, y1, fill="darkblue", font="Times 12  bold",text=text1)
        nombr_variable_str = tk.StringVar(root, value=text2)
        entry1 = Entry(root, width=ancho, textvariable=nombr_variable_str, font=('calibre', 10, 'normal'))
        self.canvas.create_window(x2, y2, window=entry1)

    def text_canvas_IntVar(self,x1,y1,text1,text2,ancho,x2,y2):
        global nombr_variable_int
        self.canvas.create_text(x1, y1, fill="darkblue", font="Times 12  bold",text=text1)
        nombr_variable_int = tk.IntVar(root, value=text2)
        entry1 = Entry(root, width=ancho, textvariable=nombr_variable_int, font=('calibre', 10, 'normal'))
        self.canvas.create_window(x2, y2, window=entry1)

    def text_canvas_DoubleVar(self, x1, y1, text1, text2, ancho, x2, y2):
        global nombr_variable_dbl
        self.canvas.create_text(x1, y1, fill="darkblue", font="Times 12  bold", text=text1)
        nombr_variable_dbl = tk.IntVar(root, value=text2)
        entry1 = Entry(root, width=ancho, textvariable=nombr_variable_dbl, font=('calibre', 10, 'normal'))
        self.canvas.create_window(x2, y2, window=entry1)


    def OpenFile(self):
        global todo,fecha,nmbr_sr,filename_xlsx,sheet_name_xlsx
    ##########################################
        while True:
            filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                       filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
            try:
                open(filename_xlsx, 'rb')
                break
            except FileNotFoundError:
                messagebox.showerror("Error","Algo salió mal al abrir el archivo")
                root.mainloop()
            if filename_xlsx !=0:
                root.mainloop()
        ##########################################
        self.text_canvas_StringVar(350, 70, "Nombre de hoja de excel:","Dat_in", 20, 600, 70)
        sheet_name_xlsx = nombr_variable_str
        self.text_canvas_IntVar(350, 90, "Numero de la columna de la fecha:","0", 20, 600, 90)
        num_col = nombr_variable_int
        self.text_canvas_IntVar(350,110, "Numero de la columna de la serie:","1", 20, 600, 110)
        num_col_ser = nombr_variable_int
        self.Next(relx=.5, rely=.4)
        sheet_name_xlsx = sheet_name_xlsx.get()
        num_col = num_col.get()
        num_col_ser = num_col_ser.get()
        Dat = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        todo = Dat.iloc[:]
        fecha=Dat.iloc[:, num_col]
        nmbr_sr=Dat.iloc[:, num_col_ser]
        ##########################################
        self.canvas.delete('all')
        self.imagecenit()




    def positiongraf(self,rutasafe,x,y,gridrow,gridcolumn):
        self.image = Image.open(rutasafe)
        self.photo = ImageTk.PhotoImage(self.image)
        self.c_image = self.canvas.create_image(x, y, anchor='nw', image=self.photo)
        self.canvas.grid(row=gridrow,column=gridcolumn)



    def plotserieinicial(self):
        ##########################################
        try:
            nmbr_sr
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.canvas.create_text(400, 10, fill="darkblue", font="Times 12 bold",text="")
        filename = filedialog.asksaveasfile(title="Select file",filetypes=(("png files","*.png"), ("all files", "*.*")))
        filename = filename.name
        ##################################################
        self.text_canvas_StringVar(300, 50,"Nombre del eje X:","Meses", 20, 600, 50)
        mbr_x = nombr_variable_str
        self.text_canvas_StringVar(300, 70, "Nombre del eje Y:","Índice", 20, 600, 70)
        mbr_y = nombr_variable_str
        self.text_canvas_StringVar(300, 90, "Título de la gráfica:","TSM1+2", 20, 600, 90)
        mbr_ttl = nombr_variable_str
        self.Next(relx=.5, rely=.4)
        mbr_x = mbr_x.get()
        mbr_y = mbr_y.get()
        mbr_ttl = mbr_ttl.get()
        ##################################################
        plt.plot(fecha, nmbr_sr, 'b', linewidth=0.4)
        plt.xlabel(mbr_x, fontsize=10)
        plt.ylabel(mbr_y, fontsize=10)
        plt.title(mbr_ttl)
        plt.grid(True)
        plt.savefig(filename, dpi=100)
        plt.close()
        ##################################################
        self.positiongraf(filename, 192, 48, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()
        ##################################################


    def analisisespectral(self):
        ##########################################
        try:
            nmbr_sr
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.text_canvas_StringVar(x1=300, y1=50, text1="Nombre del eje X:",
                                text2="Frecuencia (mes^-1)", ancho=30, x2=600, y2=50)
        mbr_x = nombr_variable_str
        self.text_canvas_StringVar(x1=300, y1=70, text1="Nombre del eje Y:",
                                text2="Densidad Espectral", ancho=30, x2=600, y2=70)
        mbr_y = nombr_variable_str
        self.text_canvas_StringVar(x1=300, y1=90, text1="Nombre del Titulo:",
                                text2="Densidad Espectral de SOI", ancho=30, x2=600, y2=90)
        mbr_ttl = nombr_variable_str
        self.Next(relx=.5, rely=.4)
        mbr_x = mbr_x.get()
        mbr_y = mbr_y.get()
        mbr_ttl = mbr_ttl.get()
        freqs, psd = signal.welch(nmbr_sr)
        plt.plot(freqs, psd)
        plt.title(mbr_ttl, fontsize=12)
        plt.xlabel(mbr_x, fontsize=10)
        plt.ylabel(mbr_y, fontsize=10)
        plt.show()
        plt.close()


    def filtro(self):
        ##########################################
        try:
            nmbr_sr
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        t = np.linspace(0, len(nmbr_sr), len(nmbr_sr))
        def butter_bandpass_filter(data, lowcut, highcut, fs, order=1):
            nyq = 0.5 * fs
            low = lowcut / nyq
            high = highcut / nyq
            b, a = butter(order, [low, high], btype='band')
            y = lfilter(b, a, data)
            return y
        ###############################################
        ###############################################
        self.text_canvas_StringVar(250,150,"La frecuencia del muestreo del sistema digital (fs):","1",15,550,150)
        fs = nombr_variable_str
        self.text_canvas_IntVar(235,170,"Ecriba 12 si datos son mensuales y 1 son diarios:","12",15,510,170)
        cantmes = nombr_variable_int
        self.text_canvas_StringVar(190,210,"Escriba cuasi ciclos determinados:","0.0039,0.0159,0.0239,0.0392,0.055",40,500,210)
        cuasic = nombr_variable_str
        self.text_canvas_StringVar(120,250,"Escriba lowcut:","0.0003,0.008,0.0197,0.035,0.05",40,350,250)
        lowcut = nombr_variable_str
        self.text_canvas_StringVar(120,290,"Escriba highcut:","0.008,0.0197,0.035,0.043,0.068",40,350,290)
        highcut = nombr_variable_str
        self.text_canvas_StringVar(340,350,"Escriba el nombre de la serie:","TSM1+2_",20,580,350)
        nombre_serie = nombr_variable_str
        self.text_canvas_StringVar(330,390,"Escriba el nombre de la Hoja de excel:","TSM1+2", 20, 580,390)
        nombre_hoja = nombr_variable_str
        self.text_canvas_StringVar(330, 430,"Escriba el nombre de la eje X:","Meses",20,580,430)
        nombre_X = nombr_variable_str
        self.text_canvas_StringVar(330,470,"Escriba el nombre de la eje Y:","Anomalia",20,580,470)
        nombre_Y = nombr_variable_str
        self.canvas.create_text(340, 510, fill="darkblue", font="Times 12  bold",
                                text="Los componentes filtrados se grabaran en el archivo inicial")
        self.Next(relx=.5, rely=.1)
        fs = fs.get()
        cantmes = cantmes.get()
        cuasic = cuasic.get()
        lowcut = lowcut.get()
        highcut = highcut.get()
        nombre_serie = nombre_serie.get()
        nombre_hoja = nombre_hoja.get()
        nombre_X = nombre_X.get()
        nombre_Y = nombre_Y.get()
        fs = np.fromstring(fs, dtype=float, sep=',')
        cuasic = np.fromstring(cuasic, dtype=float, sep=',')
        lowcut = np.fromstring(lowcut, dtype=float, sep=',')
        highcut = np.fromstring(highcut, dtype=float, sep=',')
        ###############################################
        ExcelWorkbook = load_workbook(filename_xlsx)
        writer = pd.ExcelWriter(filename_xlsx, engine='openpyxl')
        writer.book = ExcelWorkbook
        y = {}
        df = pd.DataFrame()
        df1 = pd.DataFrame(fecha.dt.strftime('%Y-%m'))
        df1.columns = ['Date']
        df1.to_excel(writer, sheet_name=nombre_hoja, index=False, engine='xlsxwriter')
        df2 = pd.DataFrame(round(nmbr_sr,2))
        df2.to_excel(writer, sheet_name=nombre_hoja, startcol=1, index=False, engine='xlsxwriter')
        for j in range (len(lowcut)):
            y[j] = butter_bandpass_filter(nmbr_sr, lowcut[j], highcut[j], fs)
            nomb = str(np.round((1/cuasic[j])/cantmes,1))
            df[nombre_serie + nomb] =  y[j]
            df = pd.DataFrame(data=df)
            df.to_excel(writer, sheet_name=nombre_hoja, startcol=2, index=False, engine='xlsxwriter')
        writer.save()
        writer.close()
        plt.plot(t, nmbr_sr, label=nombre_serie, linewidth=1)
        for j in range (len(lowcut)):
            y[j] = butter_bandpass_filter(nmbr_sr, lowcut[j], highcut[j], fs)
            nomb = str(np.round((1/cuasic[j])/cantmes,1))
            plt.plot(t, y[j], label=nombre_serie + nomb, linewidth=0.5)
        plt.xlabel(nombre_X, fontsize=10)
        plt.ylabel(nombre_Y, fontsize=10)
        plt.legend(loc='best')
        plt.grid(True)
        rutasafe = ('Figuras' + os.sep + nombre_serie + 'ComponentesFiltrados.png')
        plt.savefig(rutasafe, dpi=100)
        plt.close()
        ##################################################
        self.positiongraf(rutasafe, 250, 150, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        ##################################################
        self.imagecenit()
        ##################################################





    def crosscorrel(self):
        ###############################################
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        ##########################################
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.text_canvas_StringVar(320,100,"Nombre de la hoja del excel de la 1 serie:","Dat_in",20,610,100)
        sheet_name_xlsx1 = nombr_variable_str
        self.text_canvas_IntVar(340,120,"Numero de la columna de la fecha:","0",20,610,120)
        num_col = nombr_variable_int
        self.text_canvas_IntVar(350,140,"Numero de la columna de la serie 1:","1",20,610,140)
        num_col_ser1 = nombr_variable_int
        self.text_canvas_StringVar(320, 160, "Nombre de la hoja del excel de la 2 serie:", "Dat_in", 20, 610, 160)
        sheet_name_xlsx2 = nombr_variable_str
        self.text_canvas_IntVar(350,180,"Numero de la columna de la serie 2:","6",20,610,180)
        num_col_ser2 = nombr_variable_int
        self.Next(relx=.5, rely=.4)
        sheet_name_xlsx1 = sheet_name_xlsx1.get()
        sheet_name_xlsx2 = sheet_name_xlsx2.get()
        num_col = num_col.get()
        num_col_ser1 = num_col_ser1.get()
        num_col_ser2 = num_col_ser2.get()
        Dat1 = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx1)
        Dat2 = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx2)
        print(filename_xlsx)
        print(Dat1)
        fecha = Dat1.iloc[:, num_col]
        nmbr_sr1 = Dat1.iloc[:, num_col_ser1]
        nmbr_sr2 = Dat2.iloc[:, num_col_ser2]
        ###############################################
        time = np.linspace(0, len(nmbr_sr2), len(nmbr_sr2))
        ###############################################
        coef = plt.xcorr(nmbr_sr1, nmbr_sr2, usevlines=True, maxlags=20, normed=True, lw=2)
        if np.max(coef[1][:]) > abs(np.min(coef[1][:])):
            n = np.argwhere(coef[1][:] == np.max((coef[1][:])))
        else:
            n = np.argwhere(coef[1][:] == np.min((coef[1][:])))
        coefmax = float(coef[1][n])
        indexmax = float(coef[0][n])
        ###########################################
        self.canvas.create_text(250, 50, fill="darkblue", font="Times 12  bold",
                                text='Coeficiente de Correlción maximo=')
        self.canvas.create_text(450, 50, fill="red", font="Times 12  bold",
                                text=np.round(coefmax,2))
        self.canvas.create_text(250, 70, fill="darkblue", font="Times 12  bold",
                                text='Rezago=')
        self.canvas.create_text(450, 70, fill="red", font="Times 12  bold",
                                text=indexmax)
        self.canvas.create_text(650, 50, fill="darkblue", font="Times 12  bold",
                                text='1 serie=')
        self.canvas.create_text(750, 50, fill="red", font="Times 12  bold",
                                text=nmbr_sr1.name)
        self.canvas.create_text(650, 70, fill="darkblue", font="Times 12  bold",
                                text='2 serie=')
        self.canvas.create_text(750, 70, fill="red", font="Times 12  bold",
                                text=nmbr_sr2.name)
        ##################################################
        plt.xcorr(nmbr_sr1, nmbr_sr2, usevlines=True, maxlags=20, normed=True, lw=2)
        plt.xlabel('Rezago (mes)', fontsize=10)
        plt.ylabel('Correlación cruzada', fontsize=10)
        plt.grid(True)
        rutasafe = ('Figuras' + os.sep + 'Croscorr.png')
        plt.savefig(rutasafe, dpi=100)
        plt.close()
        ##################################################
        self.positiongraf(rutasafe, 250, 150, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()



    def correl(self):
        ###############################################
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        ##########################################
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.text_canvas_StringVar(320,100,"Nombre de la hoja del excel de la 1 serie:","TSM1+2_Ta_AMA1",20,610,100)
        sheet_name_xlsx1 = nombr_variable_str
        self.text_canvas_IntVar(350,130,"Numero de la columna de la serie 1:","0",20,610,130)
        num_col_ser1 = nombr_variable_int
        self.text_canvas_StringVar(320, 160, "Nombre de la hoja del excel de la 2 serie:", "TSM1+2_Ta_AMA1", 20, 610, 160)
        sheet_name_xlsx2 = nombr_variable_str
        self.text_canvas_IntVar(350,190,"Numero de la columna de la serie 2:","1",20,610,190)
        num_col_ser2 = nombr_variable_int
        self.Next(relx=.5, rely=.4)
        sheet_name_xlsx1 = sheet_name_xlsx1.get()
        sheet_name_xlsx2 = sheet_name_xlsx2.get()
        num_col_ser1 = num_col_ser1.get()
        num_col_ser2 = num_col_ser2.get()
        Dat1 = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx1)
        Dat2 = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx2)
        print(filename_xlsx)
        print(Dat1)
        nmbr_sr1 = Dat1.iloc[:, num_col_ser1]
        nmbr_sr2 = Dat2.iloc[:, num_col_ser2]
        ###############################################
        time = np.linspace(0, len(nmbr_sr2), len(nmbr_sr2))
        ###############################################
        coef, p = scipy.stats.pearsonr(nmbr_sr1, nmbr_sr2)  # Pearson's r
        ###########################################
        self.canvas.create_text(250, 50, fill="darkblue", font="Times 12  bold",
                                text='Coeficiente de Correlción=')
        self.canvas.create_text(450, 50, fill="red", font="Times 12  bold",
                                text=np.round(coef,2))
        self.canvas.create_text(250, 70, fill="darkblue", font="Times 12  bold",
                                text='p=')
        self.canvas.create_text(450, 70, fill="red", font="Times 12  bold",
                                text=np.round(p,4))
        self.canvas.create_text(650, 50, fill="darkblue", font="Times 12  bold",
                                text='1 serie=')
        self.canvas.create_text(750, 50, fill="red", font="Times 12  bold",
                                text=nmbr_sr1.name)
        self.canvas.create_text(650, 70, fill="darkblue", font="Times 12  bold",
                                text='2 serie=')
        self.canvas.create_text(750, 70, fill="red", font="Times 12  bold",
                                text=nmbr_sr2.name)
        ##################################################
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()






    def analisis_regresion(self):
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        ##########################################
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.text_canvas_StringVar(350,70,"Nombre de hoja de excel de Y y X:","SOI",20,620,70)
        sheet_name_xlsx = nombr_variable_str
        self.text_canvas_IntVar(350,90,"Numero de la columna de la fecha:","0",20,620,90)
        num_col_t = nombr_variable_int
        self.text_canvas_IntVar(360,110,"Numero de la columna de la serie Y:","1",20,620,110)
        num_col_ser_y = nombr_variable_int
        self.text_canvas_IntVar(310,130,"Numero de componentes:","2",20,620,130)
        num_comp = nombr_variable_int
        self.Next(relx=.5, rely=.4)
        num_comp = num_comp.get()
        num_col_ser_x = [0]*num_comp
        for i in range(num_comp):
            self.canvas.create_text(350, 150, fill="darkblue", font="Times 12  bold",
                                    text="Numero de la columna de la serie X" + str(i))
            num_col_ser_x[i] = tk.IntVar(root, value=str(i+1)) # "2"
            entry1 = Entry(root, width=20, textvariable=num_col_ser_x[i], font=('calibre', 10, 'normal'))
            self.canvas.create_window(610, 150, window=entry1)
            self.Next(relx=.5, rely=.4)
        #############################
        self.Next(relx=.5, rely=.4)
        sheet_name_xlsx = sheet_name_xlsx.get()
        num_col_t = num_col_t.get()
        num_col_ser_y = num_col_ser_y.get()
        Dat_yx = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        fecha = Dat_yx.iloc[:, num_col_t]
        nmbr_sr_y = Dat_yx.iloc[:, num_col_ser_y]
        nmbr_sr_x = [0.0]*num_comp
        for i in range(num_comp):
            num_col_ser_x[i] = num_col_ser_x[i].get()
            nmbr_sr_x[i] = Dat_yx.iloc[:, num_col_ser_x[i]]
        ####################################
        df_yx = pd.DataFrame(data=Dat_yx)
        y = (df_yx[nmbr_sr_y.name]).astype(float)
        X = df_yx[[nmbr_sr_x[i].name for i in range(num_comp)]]
        print((X))
        model = sm.OLS(endog=y, exog=X).fit()
        predictions = model.predict(X)
        ###################################
        df = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        ExcelWorkbook = load_workbook(filename_xlsx)
        writer = pd.ExcelWriter(filename_xlsx, engine='openpyxl')
        writer.book = ExcelWorkbook
        df[nmbr_sr_y.name + 'Predict'] = round(predictions,3)
        df = pd.DataFrame(data=df)
        df = df[df.filter(regex='^(?!Unnamed)').columns]
        df.to_excel(writer, sheet_name=sheet_name_xlsx, index=False, engine='xlsxwriter')
        writer.save()
        writer.close()
        ##################################################
        self.canvas.create_text(250, 50, fill="darkblue", font="Times 10  bold",
                                text='Coeficiente de regresión:')
        self.canvas.create_text(450, 90, fill="red", font="Times 10  bold",
                                text=model.params.to_string())
        self.canvas.create_text(150, 150, fill="darkblue", font="Times 10  bold",
                                text='Resultados de regresión:')
        self.canvas.create_text(450, 470, fill="red", font="Times 10  bold",
                                text=model.summary())
        ##################################################
        self.Next(relx=.5, rely=.9)
        time = np.linspace(0, len(fecha), len(fecha))
        plt.plot(time, y, 'b-', label="Data")
        plt.plot(time, predictions, 'r-.', label="Predictions")
        plt.xlabel('Meses', fontsize=10)
        plt.ylabel('Anomalia', fontsize=10)
        plt.legend(loc='best')
        plt.grid(True)
        rutasafe = ('Figuras' + os.sep + 'Regresion.png')
        plt.savefig(rutasafe, dpi=100)
        plt.close()
        ##################################################
        self.positiongraf(rutasafe, 250, 150, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()
    ##############################################################






    def analisis_parametros_ARIMA(self):
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        ##########################################
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.text_canvas_StringVar(350,70,"Nombre de la hoja de excel:","Pr_AND21",20,600,70)
        sheet_name_xlsx = nombr_variable_str
        self.text_canvas_StringVar(350,90,"Fecha de inicio de la serie:","1/1/1971",20,600,90)
        fecha_inicio = nombr_variable_str
        self.text_canvas_StringVar(350,110,"Nombre de la columna de la serie:","Pr_AND2",20,600,110)
        nombr_col_ser = nombr_variable_str
        self.canvas.create_text(350, 130, fill="darkblue", font="Helvetica 10",text="Parámetros del modelo ARIMA")
        self.text_canvas_IntVar(350, 150, "start_p:", "1", 20, 600, 150)
        n1 = nombr_variable_int
        self.text_canvas_IntVar(350, 170, "start_q:", "1", 20, 600, 170)
        n2 = nombr_variable_int
        self.text_canvas_IntVar(350, 190, "max_p:", "3", 20, 600, 190)
        n3 = nombr_variable_int
        self.text_canvas_IntVar(350, 210, "max_q:", "3", 20, 600, 210)
        n4 = nombr_variable_int
        self.text_canvas_IntVar(350, 230, "m:", "12", 20, 600, 230)
        n5 = nombr_variable_int
        self.text_canvas_IntVar(350, 250, "start_P:", "0", 20, 600, 250)
        n6 = nombr_variable_int
        self.text_canvas_StringVar(350, 270, "seasonal:", "True", 20, 600, 270)
        n7 = nombr_variable_str
        self.text_canvas_IntVar(350, 290, "d:", "0", 20, 600, 290)
        n8 = nombr_variable_int
        self.text_canvas_IntVar(350, 310, "D:", "1", 20, 600, 310)
        n9 = nombr_variable_int
        self.text_canvas_StringVar(350, 330, "trace:", "True", 20, 600, 330)
        n10 = nombr_variable_str
        self.text_canvas_StringVar(350, 350, "error_action:", "ignore", 20, 600, 350)
        n11 = nombr_variable_str
        self.text_canvas_StringVar(350, 370, "suppress_warnings:", "True", 20, 600, 370)
        n12 = nombr_variable_str
        self.text_canvas_StringVar(350, 390, "stepwise:", "True", 20, 600, 390)
        n13 = nombr_variable_str
        self.canvas.create_text(500, 570, fill="black", font="Helvetica 10",
                                text="El cálculo de los coeficientes del modelo puede tardar\n"
                                     "varios minutos, dependiendo de la capacitad del computador")
        self.Next(relx=.5, rely=.8)
        sheet_name_xlsx = sheet_name_xlsx.get()
        fecha_inicio = fecha_inicio.get()
        nombr_col_ser = nombr_col_ser.get()
        n1 = n1.get()
        n2 = n2.get()
        n3 = n3.get()
        n4 = n4.get()
        n5 = n5.get()
        n6 = n6.get()
        n7 = n7.get()
        n8 = n8.get()
        n9 = n9.get()
        n10 = n10.get()
        n11 = n11.get()
        n12 = n12.get()
        n13 = n13.get()
        Dat = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        df = pd.DataFrame(data=Dat)
        rng = pd.date_range(fecha_inicio, periods=len(df), freq='M')
        df = df.set_index(rng)
        datos_in = df[nombr_col_ser]
        ##################################################
        warnings.filterwarnings("ignore")
        stepwise_fit = auto_arima(datos_in, start_p=n1, start_q=n2,
                                  max_p=n3, max_q=n4, m=n5,
                                  start_P=n6, seasonal=n7,
                                  d=n8, D=n9, trace=n10,
                                  error_action=n11,
                                  suppress_warnings=n12,
                                  stepwise=n13);
        ##################################################
        self.canvas.create_text(250, 50, fill="darkblue", font="Times 12  bold",
                                text='Grabar los resultados de auto_arima, como por ejemplo:')
        self.canvas.create_text(250, 70, fill="darkblue", font="Times 12  bold",
                                text='Model: SARIMAX(2, 0, 3)x(2, 1, 0, 12):')
        self.canvas.create_text(250, 100, fill="darkblue", font="Times 12  bold",
                                text='Resultados de auto_arima:')
        self.canvas.create_text(450, 380, fill="red", font="Times 12  bold",
                                text=stepwise_fit.summary())
        ##################################################
        self.Next(relx=.5, rely=.9)
        ##################################################

    def test_resultados_modelo_arima(self):
        global test, train, datos_in, order_arima, seasnl_trend_arima
        global filename_xlsx, sheet_name_xlsx, nombr_col_ser
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        ##########################################
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        self.text_canvas_StringVar(350,70,"Nombre de hoja de excel:","Pr_AND21",20,580,70)
        sheet_name_xlsx = nombr_variable_str
        self.text_canvas_StringVar(350,90,"Fecha de inicio de la serie:","1/1/1971",20,600,90)
        fecha_inicio = nombr_variable_str
        self.text_canvas_StringVar(350,130,"Nombre de la columna de la serie:","Pr_AND2",20,600,130)
        nombr_col_ser = nombr_variable_str
        self.text_canvas_IntVar(350,150,"Cantidad de puntos para el test:","12",20,600,150)
        cantd_puntos = nombr_variable_int
        self.text_canvas_StringVar(350,170,"Order de ARIMA:","2, 0, 2",20,600,170)
        order_arima = nombr_variable_str
        self.text_canvas_StringVar(330,190,"Seasonal order de ARIMA:","2, 1, 0, 12",20,600,190)
        seasnl_trend_arima = nombr_variable_str
        self.Next(relx=.5, rely=.3)
        sheet_name_xlsx = sheet_name_xlsx.get()
        fecha_inicio = fecha_inicio.get()
        nombr_col_ser = nombr_col_ser.get()
        Dat = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        df = pd.DataFrame(data=Dat)
        rng = pd.date_range(fecha_inicio, periods=len(df), freq='M')
        df = df.set_index(rng)
        datos_in = df[nombr_col_ser]

        cantd_puntos = cantd_puntos.get()
        order_arima = order_arima.get()
        order_arima = np.fromstring(order_arima, dtype=float, sep=',')
        seasnl_trend_arima = seasnl_trend_arima.get()
        seasnl_trend_arima = np.fromstring(seasnl_trend_arima, dtype=float, sep=',')

        train = datos_in.iloc[:len(datos_in) - cantd_puntos]
        test = datos_in.iloc[len(datos_in) - cantd_puntos:]
        model = SARIMAX(train, order=(order_arima), seasonal_order=(seasnl_trend_arima))
        result = model.fit()
        print(result.summary())
        ##################################################
        self.canvas.create_text(450, 350, fill="red", font="Times 12  bold",
                                text=result.summary())
        ##################################################
        self.Next(relx=.5, rely=.9)
        ##################################################
        start = len(train)
        end = len(train) + len(test) - 1
        predictions = result.predict(start, end, typ='levels').rename("Predictions")
        predictions.plot(legend=True)
        test.plot(legend=True)
        plt.xlabel('Meses', fontsize=10)
        plt.ylabel('Anomalia', fontsize=10)
        plt.legend(loc='best')
        plt.grid(True)
        rutasafe = ('Figuras' + os.sep + 'Predict_ARIMA_test_un_ano.png')
        plt.savefig(rutasafe, dpi=100)
        plt.close()
        ##################################################
        self.positiongraf(rutasafe, 250, 150, 2, 1)
        ##################################################
        print(rmse(test, predictions))
        self.canvas.create_text(250, 50, fill="darkblue", font="Times 12  bold",
                                text='root_mean_squared_error =')
        self.canvas.create_text(450, 50, fill="red", font="Times 12  bold",
                                text=round(rmse(test, predictions), 2))
        print(mean_squared_error(test, predictions))
        self.canvas.create_text(250, 70, fill="darkblue", font="Times 12  bold",
                                text='mean_squared_error =')
        self.canvas.create_text(450, 70, fill="red", font="Times 12  bold",
                                text=round(mean_squared_error(test, predictions), 2))
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()




    def mostrar_pronostico(self):
        ##########################################
        try:
            datos_in
        except:
            messagebox.showerror("Error", "Hay que hacer el test ARIMA")
            root.mainloop()
        ##########################################
        self.text_canvas_IntVar(330,70,"Cantidad de meses para predecir:","36",20,600,70)
        cantd_meses = nombr_variable_int
        self.text_canvas_StringVar(300,100,"Nombre del eje X:","Meses",20,600,100)
        mbr_x = nombr_variable_str
        self.text_canvas_StringVar(300,130,"Nombre del eje Y:","Anomalia",20,600,130)
        mbr_y = nombr_variable_str
        self.text_canvas_StringVar(300,160,"Nombre del Titulo:","Pr_AND21",20,600,160)
        mbr_ttl = nombr_variable_str
        self.Next(relx=.5, rely=.3)
        cantd_meses = cantd_meses.get()
        mbr_x = mbr_x.get()
        mbr_y = mbr_y.get()
        mbr_ttl = mbr_ttl.get()
        model = SARIMAX(datos_in, order=(order_arima), seasonal_order=(seasnl_trend_arima))
        result = model.fit()
        forecast = result.predict(start=len(datos_in), end=(len(datos_in) - 1) + cantd_meses,
                                  typ='levels').rename('Forecast')
        ###################################
        ExcelWorkbook = load_workbook(filename_xlsx)
        writer = pd.ExcelWriter(filename_xlsx, engine='openpyxl')
        writer.book = ExcelWorkbook
        df1 = pd.DataFrame(forecast.index.strftime('%Y-%m-%d'))
        df2 = pd.DataFrame(round(forecast, 4))
        df1.columns = ['Date']
        df1.to_excel(writer, sheet_name=(nombr_col_ser + 'forecast'), index=False, engine='xlsxwriter')
        df2.to_excel(writer, sheet_name=(nombr_col_ser + 'forecast'),startcol=1, index=False, engine='xlsxwriter')
        writer.save()
        writer.close()
        ##################################################
        datos_in.plot(legend=True)
        forecast.plot(legend=True)
        plt.xlabel(mbr_x, fontsize=10)
        plt.ylabel(mbr_y, fontsize=10)
        plt.legend(loc='best')
        plt.grid(True)
        plt.title(mbr_ttl)
        rutasafe = ('Figuras' + os.sep + 'Predict_ARIMA_'+str(cantd_meses)+'meses'+".png")
        plt.savefig(rutasafe, dpi=100)
        plt.close()
        ##################################################
        self.positiongraf(rutasafe, 250, 150, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()






    def graficar_todo(self):
        ##########################################
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        ##########################################
        self.text_canvas_StringVar(350, 70, "Nombre de hoja de excel de predicción:", "ICPr_AND2forecast", 20, 620, 70)
        sheet_name_xlsx = nombr_variable_str
        self.text_canvas_IntVar(350, 90, "Numero de la columna de la fecha:", "0", 20, 620, 90)
        num_col_t = nombr_variable_int
        self.text_canvas_IntVar(340, 110, "Numero de la columna de la serie inicial:", "1", 20, 620, 110)
        num_col_ser_y = nombr_variable_int
        self.text_canvas_IntVar(310, 130, "Numero de la columna de suma de componentes:", "2", 20, 620, 130)
        num_comp = nombr_variable_int
        self.text_canvas_IntVar(310, 150, "Numero de la columna de pronostico:", "3", 20, 620, 150)
        num_comp_prog = nombr_variable_int
        self.text_canvas_IntVar(310, 170, "Numero de la columna de pronostico2:", "4", 20, 620, 170)
        num_comp_prog2 = nombr_variable_int
        self.Next(relx=.5, rely=.4)
        num_col_t = num_col_t.get()
        num_col_ser_y = num_col_ser_y.get()
        num_comp = num_comp.get()
        num_comp_prog = num_comp_prog.get()
        num_comp_prog2 = num_comp_prog2.get()
        #############################
        sheet_name_xlsx = sheet_name_xlsx.get()
        Dat = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        fecha = Dat.iloc[:, num_col_t]
        nmbr_sr_y = Dat.iloc[:, num_col_ser_y]
        num_sr_comp = Dat.iloc[:, num_comp]
        nmbr_sr_prog = Dat.iloc[:, num_comp_prog]
        nmbr_sr_prog2 = Dat.iloc[:, num_comp_prog2]
        ##################################################
        self.canvas.create_text(400, 10, fill="darkblue", font="Times 12 bold", text="Graba file: SUM_COMPPredictforecast.png")
        filename = filedialog.asksaveasfile(title="Select file",
                                            filetypes=(("png files", "*.png"), ("all files", "*.*")))
        filename = filename.name
        self.Next(relx=.5, rely=.4)
        ##################################################
        plt.plot(fecha, nmbr_sr_y, 'b', label=nmbr_sr_y.name, linewidth=0.4)
        plt.plot(fecha, num_sr_comp, 'g', label=num_sr_comp.name, linewidth=0.4)
        plt.plot(fecha, nmbr_sr_prog, 'y', label=nmbr_sr_prog.name, linewidth=0.4)
        plt.plot(fecha, nmbr_sr_prog2, 'r', label=nmbr_sr_prog2.name, linewidth=0.4)
        if len(nmbr_sr_y) <= 36:
            cantidad_meses = 1
        else:
            cantidad_meses = 12
        plt.xticks(np.arange(0, len(nmbr_sr_y), cantidad_meses), (fecha[::cantidad_meses]), rotation='90', fontsize=6)
        plt.xlabel('Fecha', fontsize=10)
        plt.ylabel('Anomalia', fontsize=10)
        plt.title(sheet_name_xlsx)
        plt.gcf().subplots_adjust(bottom=0.15)
        plt.legend(loc='best')
        plt.savefig(filename, dpi=100)
        plt.close()
        self.positiongraf(filename, 250, 150, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()
        ##################################################

    def graficar_una_serie(self):
        ##########################################
        filename_xlsx = filedialog.askopenfilename(title="Select file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        try:
            open(filename_xlsx, 'rb')
        except:
            messagebox.showerror("Error", "Hay que abrir el archivo inicial")
            root.mainloop()
        ##########################################
        ##########################################
        self.text_canvas_StringVar(350, 70, "Nombre de hoja de excel de predicción:", "Ta_AND21", 20, 620, 70)
        sheet_name_xlsx = nombr_variable_str
        self.text_canvas_IntVar(350, 90, "Numero de la columna de la fecha:", "0", 20, 620, 90)
        num_col_t = nombr_variable_int
        self.text_canvas_IntVar(340, 110, "Numero de la columna de la serie inicial:", "4", 20, 620, 110)
        num_col_ser_y = nombr_variable_int
        self.text_canvas_IntVar(310, 130, "Numero de la columna de pronostico:", "9", 20, 620, 130)
        num_comp = nombr_variable_int
        self.Next(relx=.5, rely=.4)
        num_col_t = num_col_t.get()
        num_col_ser_y = num_col_ser_y.get()
        num_comp = num_comp.get()
        #############################
        sheet_name_xlsx = sheet_name_xlsx.get()
        Dat = pd.read_excel(filename_xlsx, sheet_name=sheet_name_xlsx)
        fecha = Dat.iloc[:, num_col_t]
        nmbr_sr_y = Dat.iloc[:, num_col_ser_y]
        num_sr_comp = Dat.iloc[:, num_comp]
        ##################################################
        while True:
            self.canvas.create_text(400, 10, fill="darkblue", font="Times 12 bold", text="Guardar archivo *.png")
            filename_png = filedialog.asksaveasfile(title="Select file",
                                                filetypes=(("png files", "*.png"), ("all files", "*.*")))
            try:
                if filename_png != 0:
                    break
            except FileNotFoundError:
                messagebox.showerror("Error","Algo salió mal al guardar el archivo")
                root.mainloop()
        filename = filename_png.name
        self.Next(relx=.5, rely=.4)
        ##################################################
        plt.plot(fecha, nmbr_sr_y, 'b', label=nmbr_sr_y.name, linewidth=0.4)
        plt.plot(fecha, num_sr_comp, 'r', label=num_sr_comp.name, linewidth=0.4)
        if len(nmbr_sr_y) <= 36:
            cantidad_meses = 1
        else:
            cantidad_meses = 12
        plt.xticks(np.arange(0, len(nmbr_sr_y), cantidad_meses), (fecha[::cantidad_meses]), rotation='90', fontsize=6)
        plt.xlabel('Fecha', fontsize=10)
        plt.ylabel('Anomalia', fontsize=10)
        plt.title(sheet_name_xlsx)
        plt.gcf().subplots_adjust(bottom=0.15)
        plt.legend(loc='best')
        plt.savefig(filename, dpi=100)
        plt.close()
        self.positiongraf(filename, 250, 150, 2, 1)
        ##################################################
        self.Next(relx=.5, rely=.9)
        self.imagecenit()
        ##################################################
    ##########################################################################
if __name__ == '__main__':
    root = tk.Tk()
    root.title('ASCCPA')
    root.geometry("1024x768")
    app = Window(root)
    root.mainloop()

